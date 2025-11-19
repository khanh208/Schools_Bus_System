from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from .models import Attendance  # Thêm import model Attendance

@admin.register(Attendance)
class AttendanceAdmin(admin.ModelAdmin):
    # ... code cũ ...
    
    actions = ['export_attendance_report', 'send_notification_to_parents']
    
    def export_attendance_report(self, request, queryset):
        """Xuất báo cáo điểm danh ra Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Báo cáo điểm danh"
        
        # Styling
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f'BÁO CÁO ĐIỂM DANH - {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        title_cell.font = Font(bold=True, size=14, color="2E7D32")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            'STT', 'Mã HS', 'Họ tên', 'Lớp', 'Tuyến', 
            'Loại', 'Trạng thái', 'Giờ điểm danh', 'Nhiệt độ', 'Ghi chú'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_num, att in enumerate(queryset.order_by('-check_time'), 4):
            ws.cell(row=row_num, column=1, value=row_num - 3).border = border
            ws.cell(row=row_num, column=2, value=att.student.student_code).border = border
            ws.cell(row=row_num, column=3, value=att.student.full_name).border = border
            ws.cell(row=row_num, column=4, value=att.student.class_obj.name if att.student.class_obj else '').border = border
            ws.cell(row=row_num, column=5, value=att.trip.route.route_code).border = border
            ws.cell(row=row_num, column=6, value=att.get_attendance_type_display()).border = border
            
            # Status cell with color
            status_cell = ws.cell(row=row_num, column=7, value=att.get_status_display())
            status_cell.border = border
            if att.status == 'present':
                status_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                status_cell.font = Font(color="2E7D32", bold=True)
            elif att.status == 'absent':
                status_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                status_cell.font = Font(color="C62828", bold=True)
            elif att.status == 'late':
                status_cell.fill = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")
                status_cell.font = Font(color="F57C00", bold=True)
            
            ws.cell(row=row_num, column=8, value=att.check_time.strftime('%d/%m/%Y %H:%M') if att.check_time else '').border = border
            ws.cell(row=row_num, column=9, value=f"{att.temperature}°C" if att.temperature else '').border = border
            ws.cell(row=row_num, column=10, value=att.notes or '').border = border
        
        # Summary
        summary_row = len(queryset) + 5
        ws.cell(row=summary_row, column=1, value='TỔNG KẾT:').font = Font(bold=True)
        
        total = queryset.count()
        present = queryset.filter(status='present').count()
        absent = queryset.filter(status='absent').count()
        late = queryset.filter(status='late').count()
        
        ws.cell(row=summary_row, column=3, value=f'Tổng: {total}')
        ws.cell(row=summary_row, column=4, value=f'Có mặt: {present}').font = Font(color="2E7D32")
        ws.cell(row=summary_row, column=5, value=f'Vắng: {absent}').font = Font(color="C62828")
        ws.cell(row=summary_row, column=6, value=f'Muộn: {late}').font = Font(color="F57C00")
        ws.cell(row=summary_row, column=7, value=f'Tỷ lệ: {(present/total*100):.1f}%' if total > 0 else '0%').font = Font(bold=True)
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'diem_danh_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        self.message_user(request, f'✓ Đã xuất báo cáo {queryset.count()} bản ghi điểm danh.')
        return response
    
    export_attendance_report.short_description = "📊 Xuất báo cáo Excel"
    
    def send_notification_to_parents(self, request, queryset):
        """Gửi thông báo cho phụ huynh"""
        from apps.notifications.models import Notification
        
        sent_count = 0
        for attendance in queryset:
            if not attendance.parent_notified:
                # Tạo thông báo
                Notification.objects.create(
                    user=attendance.student.parent.user,
                    title=f"Điểm danh: {attendance.student.full_name}",
                    message=f"Con em {attendance.get_status_display()} lúc {attendance.check_time.strftime('%H:%M %d/%m/%Y') if attendance.check_time else 'N/A'}",
                    notification_type='info',
                    sent_via='in_app'
                )
                attendance.parent_notified = True
                attendance.save()
                sent_count += 1
        
        self.message_user(request, f'✓ Đã gửi thông báo đến {sent_count} phụ huynh.')
    
    send_notification_to_parents.short_description = "📧 Gửi thông báo phụ huynh"